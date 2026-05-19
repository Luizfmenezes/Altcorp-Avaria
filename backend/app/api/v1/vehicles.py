from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Optional
from datetime import datetime, date, timedelta
import csv
import codecs
import io
from app.core.database import get_db
from app.core.deps import require_admin_or_analyst, require_web_access, get_current_user
from app.models.user import User
from app.models.vehicle import Vehicle
from app.models.inspection import Inspection
from app.models.damage import Damage
from app.schemas.vehicle import VehicleCreate, VehicleUpdate, VehicleOut
from app.schemas.inspection import InspectionOut, DamageOut, PhotoOut
from app.services.storage import storage

router = APIRouter()

ALLOWED_IMAGE_TYPES = {"image/jpeg", "image/png", "image/webp"}
MAX_PHOTO_BYTES = 8 * 1024 * 1024


def _serialize_vehicle(v: Vehicle) -> VehicleOut:
    return VehicleOut(
        id=v.id,
        plate=v.plate,
        prefix=v.prefix,
        model=v.model,
        chassis=v.chassis,
        year=v.year,
        vehicle_type=v.vehicle_type,
        is_active=v.is_active,
        default_photo_url=storage.public_url(v.default_photo_key) if v.default_photo_key else None,
        created_at=v.created_at,
    )


@router.get("", response_model=List[VehicleOut])
def list_vehicles(
    q: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_web_access),
):
    query = db.query(Vehicle)
    if q:
        like = f"%{q.upper()}%"
        query = query.filter(or_(Vehicle.plate.ilike(like), Vehicle.prefix.ilike(like), Vehicle.model.ilike(like)))
    return [_serialize_vehicle(v) for v in query.order_by(Vehicle.plate).limit(200).all()]


@router.post("/import-csv", response_model=dict)
def import_vehicles_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_or_analyst),
):
    """Importa veículos via CSV com cabecalho: plate,prefix,model,chassis,year,vehicle_type"""
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(400, "Envie um arquivo .csv")
    try:
        raw = file.file.read()
        text = codecs.decode(raw, "utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(422, "CSV vazio ou sem cabecalho")
        created = 0
        skipped = 0
        errors: list[str] = []
        for i, row in enumerate(reader, start=2):  # 1-based + header
            plate = (row.get("plate") or "").strip()
            if not plate:
                errors.append(f"Linha {i}: placa vazia")
                skipped += 1
                continue
            plate = plate.upper()
            if db.query(Vehicle).filter(Vehicle.plate == plate).first():
                skipped += 1
                continue
            prefix = (row.get("prefix") or "").strip() or None
            model = (row.get("model") or "").strip()
            if not model:
                errors.append(f"Linha {i}: modelo obrigatório")
                skipped += 1
                continue
            chassis = (row.get("chassis") or "").strip() or None
            year_raw = (row.get("year") or "").strip()
            year = int(year_raw) if year_raw.isdigit() else None
            vtype = (row.get("vehicle_type") or "bus").strip().lower()
            if vtype not in ("bus", "car"):
                vtype = "bus"
            v = Vehicle(
                plate=plate,
                prefix=prefix,
                model=model,
                chassis=chassis,
                year=year,
                vehicle_type=vtype,
            )
            db.add(v)
            created += 1
        db.commit()
        return {"created": created, "skipped": skipped, "errors": errors}
    except Exception as e:
        db.rollback()
        raise HTTPException(422, f"Erro ao processar CSV: {str(e)}")


@router.post("/import-xlsx", response_model=dict)
def import_vehicles_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_or_analyst),
):
    """Importa veículos via planilha Excel (.xlsx / .xls).

    Primeira linha = cabeçalho. Colunas: plate (obrigatória), prefix,
    model (obrigatória), chassis, year, vehicle_type (bus/car).
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Envie um arquivo .xlsx (Excel)")
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "Dependência openpyxl ausente no servidor")

    try:
        wb = load_workbook(io.BytesIO(file.file.read()), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            raise HTTPException(422, "Planilha vazia")

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        col = {name: idx for idx, name in enumerate(header)}
        if not {"plate", "model"}.issubset(col):
            raise HTTPException(422, "Cabeçalho deve conter as colunas 'plate' e 'model'")

        def cell(row: tuple, name: str) -> Optional[str]:
            idx = col.get(name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return None
            return str(row[idx]).strip() or None

        created = 0
        skipped = 0
        errors: list[str] = []
        for i, row in enumerate(rows[1:], start=2):  # 1-based + header
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            plate = cell(row, "plate")
            if not plate:
                errors.append(f"Linha {i}: placa vazia")
                skipped += 1
                continue
            plate = plate.upper()
            if db.query(Vehicle).filter(Vehicle.plate == plate).first():
                skipped += 1
                continue
            model = cell(row, "model")
            if not model:
                errors.append(f"Linha {i}: modelo obrigatório")
                skipped += 1
                continue
            year_raw = cell(row, "year")
            year = int(float(year_raw)) if year_raw and year_raw.replace(".", "", 1).isdigit() else None
            vtype = (cell(row, "vehicle_type") or "bus").lower()
            if vtype not in ("bus", "car"):
                vtype = "bus"
            db.add(Vehicle(
                plate=plate,
                prefix=cell(row, "prefix"),
                model=model,
                chassis=cell(row, "chassis"),
                year=year,
                vehicle_type=vtype,
            ))
            created += 1
        db.commit()
        return {"created": created, "skipped": skipped, "errors": errors[:20]}
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(422, f"Erro ao processar planilha: {str(e)}")


@router.get("/import-xlsx/template")
def import_xlsx_template(_: User = Depends(require_admin_or_analyst)):
    """Gera uma planilha .xlsx modelo para importação de veículos."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(500, "Dependência openpyxl ausente no servidor")

    wb = Workbook()
    ws = wb.active
    ws.title = "Veiculos"
    ws.append(["plate", "prefix", "model", "chassis", "year", "vehicle_type"])
    ws.append(["ABC1D23", "1003", "M.Benz OF-1721", "9BM384067VB123456", 2022, "bus"])
    ws.append(["DEF2E34", "15040", "Volvo B270F", "93S270F1234567890", 2023, "bus"])
    ws.append(["GHI3F45", "200", "Volkswagen Gol", "", 2021, "car"])
    ws.append(["JKL4G56", "", "Scania K310", "9BSC3101234567890", 2024, "bus"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="modelo_importacao_veiculos.xlsx"'},
    )


@router.post("", response_model=VehicleOut, status_code=201)
def create_vehicle(payload: VehicleCreate, db: Session = Depends(get_db), _: User = Depends(require_admin_or_analyst)):
    plate = payload.plate.upper().strip()
    if db.query(Vehicle).filter(Vehicle.plate == plate).first():
        raise HTTPException(409, "Placa já cadastrada")
    v = Vehicle(
        plate=plate,
        prefix=payload.prefix,
        model=payload.model,
        chassis=payload.chassis,
        year=payload.year,
        vehicle_type=payload.vehicle_type,
    )
    db.add(v)
    db.commit()
    db.refresh(v)
    return _serialize_vehicle(v)


@router.patch("/{vehicle_id}", response_model=VehicleOut)
def update_vehicle(vehicle_id: int, payload: VehicleUpdate, db: Session = Depends(get_db), _: User = Depends(require_admin_or_analyst)):
    v = db.get(Vehicle, vehicle_id)
    if not v:
        raise HTTPException(404, "Veículo não encontrado")
    data = payload.model_dump(exclude_unset=True)
    if "plate" in data and data["plate"]:
        plate = data["plate"].upper().strip()
        clash = db.query(Vehicle).filter(Vehicle.plate == plate, Vehicle.id != vehicle_id).first()
        if clash:
            raise HTTPException(409, "Placa já cadastrada em outro veículo")
        data["plate"] = plate
    for k, val in data.items():
        setattr(v, k, val)
    db.commit()
    db.refresh(v)
    return _serialize_vehicle(v)


@router.post("/{vehicle_id}/photo", response_model=VehicleOut)
def upload_default_photo(
    vehicle_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_or_analyst),
):
    v = db.get(Vehicle, vehicle_id)
    if not v:
        raise HTTPException(404, "Veículo não encontrado")
    if file.content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(415, "Formato inválido. Envie JPEG, PNG ou WebP.")
    content = file.file.read()
    if len(content) > MAX_PHOTO_BYTES:
        raise HTTPException(413, "Imagem maior que 8 MB.")
    key = storage.upload(io.BytesIO(content), file.content_type or "image/jpeg", prefix=f"vehicles/{vehicle_id}")
    v.default_photo_key = key
    db.commit()
    db.refresh(v)
    return _serialize_vehicle(v)


@router.get("/{vehicle_id}", response_model=VehicleOut)
def get_vehicle(vehicle_id: int, db: Session = Depends(get_db), _: User = Depends(require_web_access)):
    v = db.get(Vehicle, vehicle_id)
    if not v:
        raise HTTPException(404, "Veículo não encontrado")
    return _serialize_vehicle(v)


@router.get("/{vehicle_id}/history", response_model=List[InspectionOut])
def vehicle_history(
    vehicle_id: int,
    day: Optional[str] = Query(None, description="Filtra por dia (YYYY-MM-DD)"),
    month: Optional[str] = Query(None, description="Filtra por mês (YYYY-MM)"),
    limit: Optional[int] = Query(None, ge=1, le=200, description="Limita aos últimos N registros"),
    db: Session = Depends(get_db),
    _: User = Depends(require_web_access),
):
    v = db.get(Vehicle, vehicle_id)
    if not v:
        raise HTTPException(404, "Veículo não encontrado")

    query = db.query(Inspection).filter(Inspection.vehicle_id == vehicle_id)

    if day:
        try:
            d = date.fromisoformat(day)
        except ValueError:
            raise HTTPException(422, "Parâmetro 'day' deve estar no formato YYYY-MM-DD")
        start = datetime(d.year, d.month, d.day)
        query = query.filter(
            Inspection.performed_at >= start,
            Inspection.performed_at < start + timedelta(days=1),
        )
    elif month:
        try:
            y, m = (int(p) for p in month.split("-"))
            start = datetime(y, m, 1)
            end = datetime(y + (m // 12), (m % 12) + 1, 1)
        except (ValueError, IndexError):
            raise HTTPException(422, "Parâmetro 'month' deve estar no formato YYYY-MM")
        query = query.filter(Inspection.performed_at >= start, Inspection.performed_at < end)

    query = query.order_by(Inspection.performed_at.desc())
    if limit:
        query = query.limit(limit)

    out: List[InspectionOut] = []
    for i in query.all():
        out.append(
            InspectionOut(
                id=i.id,
                vehicle_id=i.vehicle_id,
                vehicle_plate=v.plate,
                vehicle_prefix=i.vehicle_prefix or v.prefix,
                vehicle_model=v.model,
                vehicle_type=v.vehicle_type,
                vehicle_default_photo_url=storage.public_url(v.default_photo_key) if v.default_photo_key else None,
                inspector_id=i.inspector_id,
                inspector_name=i.inspector.name if i.inspector else None,
                inspection_type=i.inspection_type,
                status=i.status,
                notes=i.notes,
                performed_at=i.performed_at,
                damages=[DamageOut.model_validate(d) for d in i.damages],
                photos=[
                    PhotoOut(
                        id=p.id,
                        object_key=p.object_key,
                        url=storage.public_url(p.object_key),
                        content_type=p.content_type,
                        created_at=p.created_at,
                    )
                    for p in i.photos
                ],
            )
        )
    return out


@router.post("/import-csv")
def import_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_admin_or_analyst),
):
    """Importa veículos via CSV. Espera colunas: placa, prefixo, modelo, chassi, ano, tipo, ativo.

    O CSV pode usar separador vírgula ou ponto-e-vírgula, com ou sem cabeçalho.
    Colunas obrigatórias: placa, modelo (ou `modelo`).
    """
    if not file.filename or not any(file.filename.lower().endswith(ext) for ext in (".csv", ".txt")):
        raise HTTPException(400, "Envie um arquivo .csv ou .txt")

    raw = file.file.read()
    # Tenta decodificar UTF-8 (com ou sem BOM), fallback para latin-1
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("latin-1")

    # Detecta separador: conta quem aparece mais na primeira linha
    first_line = text.split("\n")[0] if text else ""
    sep = ";" if first_line.count(";") > first_line.count(",") else ","

    reader = csv.reader(io.StringIO(text), delimiter=sep)
    rows = list(reader)
    if not rows:
        raise HTTPException(400, "Arquivo vazio")

    # Detecta se a primeira linha é cabeçalho
    header = [h.strip().lower() for h in rows[0]]
    has_header = any(
        h in header for h in ("placa", "modelo", "plate", "model", "prefixo", "prefix", "ano", "year")
    )

    if has_header:
        # Mapeia colunas por nome
        col_map = {h: i for i, h in enumerate(header)}
        data_rows = rows[1:]
    else:
        # Usa mapeamento posicional fixo: placa, prefixo, modelo, chassi, ano, tipo, ativo
        col_map = {
            "placa": 0, "prefixo": 1, "modelo": 2, "chassi": 3, "ano": 4, "tipo": 5, "ativo": 6,
        }
        data_rows = rows

    def get_col(row: list[str], *names: str) -> Optional[str]:
        for n in names:
            idx = col_map.get(n)
            if idx is not None and idx < len(row):
                val = row[idx].strip()
                if val:
                    return val
        return None

    created = 0
    skipped = 0
    errors: list[str] = []

    for i, row in enumerate(data_rows, start=1):
        if not row or all(c.strip() == "" for c in row):
            continue
        plate = get_col(row, "placa", "plate")
        model = get_col(row, "modelo", "model")
        if not plate or not model:
            skipped += 1
            errors.append(f"Linha {i}: placa ou modelo ausente")
            continue

        plate = plate.upper().strip()
        try:
            prefix = get_col(row, "prefixo", "prefix")
            chassis = get_col(row, "chassi", "chassis")
            year_str = get_col(row, "ano", "year")
            year = int(year_str) if year_str and year_str.isdigit() else None
            vtype = get_col(row, "tipo", "tipo_veiculo", "vehicle_type", "type")
            if vtype:
                vt = vtype.lower()
                if vt in ("carro", "car", "c"):
                    vtype = "car"
                elif vt in ("onibus", "ônibus", "bus", "b", "o"):
                    vtype = "bus"
                else:
                    vtype = "bus"
            else:
                vtype = "bus"
            active_str = get_col(row, "ativo", "active", "is_active")
            is_active = True
            if active_str and active_str.lower() in ("n", "não", "nao", "no", "false", "0", "inativo", "inativa"):
                is_active = False

            # Verifica se placa já existe
            existing = db.query(Vehicle).filter(Vehicle.plate == plate).first()
            if existing:
                # Atualiza campos
                existing.prefix = prefix or existing.prefix
                existing.model = model
                existing.chassis = chassis or existing.chassis
                existing.year = year if year is not None else existing.year
                existing.vehicle_type = vtype
                existing.is_active = is_active
                skipped += 1
            else:
                v = Vehicle(
                    plate=plate,
                    prefix=prefix,
                    model=model,
                    chassis=chassis,
                    year=year,
                    vehicle_type=vtype,
                    is_active=is_active,
                )
                db.add(v)
                created += 1
        except Exception as e:
            skipped += 1
            errors.append(f"Linha {i} ({plate}): {e}")

    db.commit()

    return {
        "ok": True,
        "created": created,
        "updated": skipped - len(errors) if skipped > len(errors) else 0,
        "skipped": len(errors),
        "errors": errors[:20],
    }


@router.get("/by-plate/{plate}", response_model=VehicleOut)
def by_plate(plate: str, db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    v = db.query(Vehicle).filter(Vehicle.plate == plate.upper().strip()).first()
    if not v:
        raise HTTPException(404, "Placa não encontrada")
    return _serialize_vehicle(v)
